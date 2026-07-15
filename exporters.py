import csv
import os
import re
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Image,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

HEADER_FILL = "1F4E78"
HEADER_FONT_COLOR = "FFFFFF"
LOGO_PATH = "logo.png"
MAX_PDF_COLUMNS = 8
MAX_PDF_CELL_CHARS = 300


def _read_rows(csv_path):
    with open(csv_path, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _safe_sheet_name(title):
    name = re.sub(r"[\\/*?:\[\]]", "", title)[:31]
    return name or "Data"


def export_excel(csv_path, xlsx_path, title="Data"):
    if not os.path.exists(csv_path):
        print(f"Hittar inte {csv_path}, hoppar över Excel-export.")
        return None

    df = pd.read_csv(csv_path, dtype=str).fillna("")
    sheet_name = _safe_sheet_name(title)
    df.to_excel(xlsx_path, index=False, sheet_name=sheet_name)

    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]

    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    header_font = Font(color=HEADER_FONT_COLOR, bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) for c in column_cells if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    wb.save(xlsx_path)
    print(f"Excel-export klar: {xlsx_path} ({len(df)} rader)")
    return xlsx_path


class _NumberedCanvas:
    def __init__(self, title):
        self.title = title

    def __call__(self, canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.drawRightString(doc.pagesize[0] - 1.5 * cm, 1 * cm, f"Sida {doc.page}")
        canvas.drawString(1.5 * cm, 1 * cm, self.title)
        canvas.restoreState()


def export_pdf(csv_path, pdf_path, title="Data"):
    if not os.path.exists(csv_path):
        print(f"Hittar inte {csv_path}, hoppar över PDF-export.")
        return None

    rows = _read_rows(csv_path)
    if not rows:
        print("Ingen data att exportera till PDF.")
        return None

    all_columns = list(rows[0].keys())
    if len(all_columns) <= MAX_PDF_COLUMNS or "Source URL" not in all_columns:
        columns = all_columns[:MAX_PDF_COLUMNS]
    else:
        columns = all_columns[: MAX_PDF_COLUMNS - 1] + ["Source URL"]

    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7, leading=9)
    header_style = ParagraphStyle(
        "header", parent=styles["Normal"], fontSize=8, leading=10,
        textColor=colors.white, fontName="Helvetica-Bold",
    )

    doc = SimpleDocTemplate(
        pdf_path, pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )

    story = []

    if os.path.exists(LOGO_PATH):
        story.append(Image(LOGO_PATH, width=4 * cm, height=2 * cm))
        story.append(Spacer(1, 0.5 * cm))

    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=16)
    story.append(Paragraph(title, title_style))
    story.append(Paragraph(
        f"Genererad {datetime.now().strftime('%Y-%m-%d %H:%M')} &ndash; {len(rows)} rader",
        styles["Normal"],
    ))
    story.append(Spacer(1, 0.5 * cm))

    def _cell_text(value):
        text = str(value)
        if len(text) > MAX_PDF_CELL_CHARS:
            text = text[:MAX_PDF_CELL_CHARS] + "..."
        return text

    table_data = [[Paragraph(col, header_style) for col in columns]]
    for row in rows:
        table_data.append([Paragraph(_cell_text(row.get(col, "")), cell_style) for col in columns])

    usable_width = landscape(A4)[0] - 3 * cm
    col_widths = [usable_width / len(columns)] * len(columns)
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{HEADER_FILL}")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    story.append(table)

    numbered = _NumberedCanvas(title)
    doc.build(story, onFirstPage=numbered, onLaterPages=numbered)
    print(f"PDF-export klar: {pdf_path} ({len(rows)} rader, {len(columns)} kolumner)")
    return pdf_path
